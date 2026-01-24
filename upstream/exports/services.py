"""
Export service for Upstream data.

Provides Excel and CSV export functionality for core data types.
"""

from typing import Optional, Dict, Any
import os
import csv
from datetime import datetime
from io import BytesIO
from django.conf import settings
from django.contrib.auth.models import User
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from upstream.models import DriftEvent, ReportRun, Customer
from upstream.alerts.models import AlertEvent, Alert
from upstream.core.models import DomainAuditEvent


class ExportService:
    """Service for exporting Upstream data to various formats."""
    
    def __init__(self, customer: Customer) -> None:
        """
        Initialize export service for a specific customer.

        Args:
            customer: Customer instance to scope exports to
        """
        self.customer = customer
    
    def export_drift_events_excel(self, report_run: ReportRun, filters: Optional[Dict[str, Any]] = None) -> BytesIO:
        """
        Export drift events to Excel format.

        Args:
            report_run: ReportRun instance
            filters: Optional dict with filters (min_severity, payer, etc.)

        Returns:
            BytesIO: Excel file as bytes
        """
        if filters is None:
            filters = {}
        
        # Get drift events
        drift_events = DriftEvent.objects.filter(
            report_run=report_run,
            customer=self.customer
        ).order_by('-severity', 'payer')
        
        # Apply filters
        if filters.get('min_severity'):
            drift_events = drift_events.filter(severity__gte=float(filters['min_severity']))
        if filters.get('payer'):
            drift_events = drift_events.filter(payer__icontains=filters['payer'])
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Drift Events"
        
        # Add headers
        headers = [
            'Payer', 'CPT Group', 'Drift Type', 'Baseline Value', 'Current Value',
            'Delta', 'Severity', 'Confidence', 'Baseline Start', 'Baseline End',
            'Current Start', 'Current End'
        ]
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Add data rows
        for event in drift_events:
            ws.append([
                event.payer,
                event.cpt_group,
                event.drift_type,
                round(event.baseline_value, 2),
                round(event.current_value, 2),
                round(event.delta_value, 2),
                round(event.severity, 2),
                round(event.confidence, 2),
                str(event.baseline_start),
                str(event.baseline_end),
                str(event.current_start),
                str(event.current_end),
            ])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    def export_alert_events_excel(self, filters: Optional[Dict[str, Any]] = None) -> BytesIO:
        """
        Export alert events to Excel format.

        Args:
            filters: Optional dict with filters (severity, date_range, etc.)

        Returns:
            BytesIO: Excel file as bytes
        """
        if filters is None:
            filters = {}
        
        # Get alert events
        alert_events = AlertEvent.objects.filter(
            customer=self.customer
        ).select_related('alert_rule').order_by('-created_at')
        
        # Apply filters
        if filters.get('severity'):
            alert_events = alert_events.filter(severity=filters['severity'])
        if filters.get('start_date'):
            alert_events = alert_events.filter(created_at__gte=filters['start_date'])
        if filters.get('end_date'):
            alert_events = alert_events.filter(created_at__lte=filters['end_date'])
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Alert Events"
        
        # Add headers
        headers = [
            'Triggered At', 'Alert Rule', 'Status', 'Severity',
            'Payer', 'CPT Group', 'Error Message'
        ]
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Add data rows
        for event in alert_events:
            ws.append([
                event.triggered_at.strftime('%Y-%m-%d %H:%M:%S'),
                event.alert_rule.name if event.alert_rule else 'N/A',
                event.status,
                event.payload.get('severity', 'N/A'),
                event.payload.get('payer', 'N/A'),
                event.payload.get('cpt_group', 'N/A'),
                event.error_message or '',
            ])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    def export_weekly_summary_excel(self, report_run: ReportRun) -> BytesIO:
        """
        Export weekly summary to Excel with multiple sheets.

        Args:
            report_run: ReportRun instance

        Returns:
            BytesIO: Excel file as bytes
        """
        # Create workbook
        wb = Workbook()
        
        # Sheet 1: Summary
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        drift_events = DriftEvent.objects.filter(
            report_run=report_run,
            customer=self.customer
        )
        
        total_events = drift_events.count()
        high_severity = drift_events.filter(severity__gte=0.7).count()
        medium_severity = drift_events.filter(severity__gte=0.4, severity__lt=0.7).count()
        low_severity = drift_events.filter(severity__lt=0.4).count()
        unique_payers = drift_events.values_list('payer', flat=True).distinct().count()
        
        # Add summary data
        ws_summary.append(['Weekly Drift Report Summary'])
        ws_summary.append([])
        ws_summary.append(['Customer', self.customer.name])
        ws_summary.append(['Report Date', datetime.now().strftime('%Y-%m-%d')])
        ws_summary.append(['Baseline Period', f"{report_run.summary_json.get('baseline_start', 'N/A')} to {report_run.summary_json.get('baseline_end', 'N/A')}"])
        ws_summary.append(['Current Period', f"{report_run.summary_json.get('current_start', 'N/A')} to {report_run.summary_json.get('current_end', 'N/A')}"])
        ws_summary.append([])
        ws_summary.append(['Total Drift Events', total_events])
        ws_summary.append(['High Severity', high_severity])
        ws_summary.append(['Medium Severity', medium_severity])
        ws_summary.append(['Low Severity', low_severity])
        ws_summary.append(['Unique Payers Affected', unique_payers])
        
        # Style summary sheet
        ws_summary['A1'].font = Font(size=14, bold=True)
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 40
        
        # Sheet 2: Drift Events (reuse existing logic)
        ws_details = wb.create_sheet(title="Drift Events")
        
        headers = [
            'Payer', 'CPT Group', 'Drift Type', 'Baseline Value', 'Current Value',
            'Delta', 'Severity', 'Confidence'
        ]
        ws_details.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws_details[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Add drift events
        for event in drift_events.order_by('-severity'):
            ws_details.append([
                event.payer,
                event.cpt_group,
                event.drift_type,
                round(event.baseline_value, 2),
                round(event.current_value, 2),
                round(event.delta_value, 2),
                round(event.severity, 2),
                round(event.confidence, 2),
            ])
        
        # Auto-adjust column widths
        for column in ws_details.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws_details.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    def log_export(self, user: User, export_type: str, record_count: int) -> None:
        """
        Log export action for audit trail.

        Args:
            user: User who performed the export
            export_type: Type of export (drift_events, alert_events, etc.)
            record_count: Number of records exported
        """
        DomainAuditEvent.objects.create(
            customer=self.customer,
            action='report_exported',
            entity_type='export',
            entity_id=export_type,
            user=user,
            metadata={
                'export_type': export_type,
                'record_count': record_count,
                'timestamp': datetime.now().isoformat()
            }
        )
