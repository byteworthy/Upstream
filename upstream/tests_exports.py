"""
Tests for export functionality.
"""

from django.test import TestCase
from django.contrib.auth.models import User
from upstream.models import Customer, UserProfile, ReportRun, DriftEvent
from upstream.alerts.models import AlertRule, AlertEvent
from upstream.exports.services import ExportService
from datetime import datetime, timedelta
from django.utils import timezone
from io import BytesIO
from openpyxl import load_workbook


class ExportServiceTests(TestCase):
    def setUp(self):
        self.customer = Customer.objects.create(name='Test Customer')
        self.user = User.objects.create_user(username='testuser', password='pass')
        UserProfile.objects.create(user=self.user, customer=self.customer, role='owner')
        
        # Create report run
        self.report_run = ReportRun.objects.create(
            customer=self.customer,
            run_type='weekly',
            status='success',
            summary_json={
                'baseline_start': '2024-01-01',
                'baseline_end': '2024-01-31',
                'current_start': '2024-02-01',
                'current_end': '2024-02-28'
            }
        )
        
        # Create drift events
        for i in range(5):
            DriftEvent.objects.create(
                customer=self.customer,
                report_run=self.report_run,
                payer=f'Payer{i}',
                cpt_group='EVAL',
                drift_type='DENIAL_RATE',
                baseline_value=0.2,
                current_value=0.6,
                delta_value=0.4,
                severity=0.8 - (i * 0.1),
                confidence=0.9,
                baseline_start=datetime(2024, 1, 1).date(),
                baseline_end=datetime(2024, 1, 31).date(),
                current_start=datetime(2024, 2, 1).date(),
                current_end=datetime(2024, 2, 28).date()
            )
        
        # Create alert rule and events
        self.alert_rule = AlertRule.objects.create(
            customer=self.customer,
            name='Test Alert',
            threshold_value=0.5
        )
        
        for i in range(3):
            AlertEvent.objects.create(
                customer=self.customer,
                alert_rule=self.alert_rule,
                status='sent',
                payload={'payer': f'Payer{i}', 'cpt_group': 'EVAL', 'severity': 0.8}
            )
    
    def test_export_service_initialization(self):
        """Test that export service initializes correctly."""
        export_service = ExportService(self.customer)
        self.assertEqual(export_service.customer, self.customer)
    
    def test_export_drift_events_excel(self):
        """Test exporting drift events to Excel."""
        export_service = ExportService(self.customer)
        excel_file = export_service.export_drift_events_excel(self.report_run)
        
        # Verify it's a BytesIO object
        self.assertIsInstance(excel_file, BytesIO)
        
        # Load workbook and verify structure
        wb = load_workbook(excel_file)
        ws = wb.active
        
        # Check headers
        headers = [cell.value for cell in ws[1]]
        self.assertIn('Payer', headers)
        self.assertIn('CPT Group', headers)
        self.assertIn('Severity', headers)
        
        # Check data rows (5 events + 1 header)
        self.assertEqual(ws.max_row, 6)
    
    def test_export_drift_events_with_filters(self):
        """Test exporting drift events with filters."""
        export_service = ExportService(self.customer)
        
        # Filter by min_severity
        excel_file = export_service.export_drift_events_excel(
            self.report_run,
            filters={'min_severity': 0.7}
        )
        
        wb = load_workbook(excel_file)
        ws = wb.active
        
        # Should have fewer rows (only high severity)
        self.assertLess(ws.max_row, 6)
    
    def test_export_alert_events_excel(self):
        """Test exporting alert events to Excel."""
        export_service = ExportService(self.customer)
        excel_file = export_service.export_alert_events_excel()
        
        # Verify it's a BytesIO object
        self.assertIsInstance(excel_file, BytesIO)
        
        # Load workbook and verify structure
        wb = load_workbook(excel_file)
        ws = wb.active
        
        # Check headers
        headers = [cell.value for cell in ws[1]]
        self.assertIn('Alert Rule', headers)
        self.assertIn('Severity', headers)
        self.assertIn('Status', headers)
        
        # Check data rows (3 events + 1 header)
        self.assertEqual(ws.max_row, 4)
    
    def test_export_weekly_summary_excel(self):
        """Test exporting weekly summary to Excel with multiple sheets."""
        export_service = ExportService(self.customer)
        excel_file = export_service.export_weekly_summary_excel(self.report_run)
        
        # Verify it's a BytesIO object
        self.assertIsInstance(excel_file, BytesIO)
        
        # Load workbook and verify structure
        wb = load_workbook(excel_file)
        
        # Should have 2 sheets
        self.assertIn('Summary', wb.sheetnames)
        self.assertIn('Drift Events', wb.sheetnames)
        
        # Check summary sheet
        ws_summary = wb['Summary']
        self.assertEqual(ws_summary['A1'].value, 'Weekly Drift Report Summary')
        
        # Check drift events sheet
        ws_details = wb['Drift Events']
        headers = [cell.value for cell in ws_details[1]]
        self.assertIn('Payer', headers)
    
    def test_export_empty_dataset(self):
        """Test exporting when no data exists."""
        # Create new customer with no data
        empty_customer = Customer.objects.create(name='Empty Customer')
        export_service = ExportService(empty_customer)
        
        # Create empty report run
        empty_report = ReportRun.objects.create(
            customer=empty_customer,
            run_type='weekly',
            status='success',
            summary_json={}
        )
        
        excel_file = export_service.export_drift_events_excel(empty_report)
        
        # Should still create valid Excel file with just headers
        wb = load_workbook(excel_file)
        ws = wb.active
        self.assertEqual(ws.max_row, 1)  # Only header row
    
    def test_log_export_creates_audit_event(self):
        """Test that log_export creates an audit trail."""
        from upstream.core.models import DomainAuditEvent
        
        export_service = ExportService(self.customer)
        export_service.log_export(self.user, 'drift_events', 5)
        
        # Verify audit event was created
        audit_events = DomainAuditEvent.objects.filter(
            customer=self.customer,
            action='report_exported',
            user=self.user
        )
        
        self.assertEqual(audit_events.count(), 1)
        audit_event = audit_events.first()
        self.assertEqual(audit_event.metadata['export_type'], 'drift_events')
        self.assertEqual(audit_event.metadata['record_count'], 5)
    
    def test_customer_isolation(self):
        """Test that exports are scoped to customer."""
        # Create second customer with data
        customer2 = Customer.objects.create(name='Customer 2')
        report_run2 = ReportRun.objects.create(
            customer=customer2,
            run_type='weekly',
            status='success',
            summary_json={}
        )
        
        DriftEvent.objects.create(
            customer=customer2,
            report_run=report_run2,
            payer='Other Payer',
            cpt_group='SURG',
            drift_type='DECISION_TIME',
            baseline_value=5.0,
            current_value=10.0,
            delta_value=5.0,
            severity=0.7,
            confidence=0.8,
            baseline_start=datetime(2024, 1, 1).date(),
            baseline_end=datetime(2024, 1, 31).date(),
            current_start=datetime(2024, 2, 1).date(),
            current_end=datetime(2024, 2, 28).date()
        )
        
        # Export for customer1 should not include customer2 data
        export_service = ExportService(self.customer)
        excel_file = export_service.export_drift_events_excel(self.report_run)
        
        wb = load_workbook(excel_file)
        ws = wb.active
        
        # Should only have customer1's 5 events
        self.assertEqual(ws.max_row, 6)
        
        # Verify no customer2 data
        for row in ws.iter_rows(min_row=2, values_only=True):
            self.assertNotEqual(row[0], 'Other Payer')
